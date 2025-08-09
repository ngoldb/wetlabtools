"""
This module contains code to parse data from the Tecan Excel sheets
"""

import string
import openpyxl
import pandas as pd

def parse_header(excel_sheet, plate_format: int):
        
        def row2list(row):
            row_list = [cell.value for cell in row]
            row_list =  [value for value in row_list if value != None and value != '' and value != ' ']
            return row_list
        
        def process_actions(iterator):
            '''Parse action list from Tecan file'''
            actions = {}
            row = next(iterator)
            row_list = row2list(row)

            while row_list != []:
                if len(row_list) == 1:
                    actions[row_list[0]] = dict()
                else:
                    actions[row_list[0]] = {'label': row_list[1]}
                row = next(iterator)
                row_list = row2list(row)
            
            for action in actions:
                row = next(iterator)
                row_list = row2list(row)
                while row_list != []:
                    label_name = row_list[0]
                    label_value = ' '.join([str(x) for x in row_list[1:]])

                    # plate area information might be stored in two rows because Tecan SUCKS!!!!!
                    if label_name == 'Plate area' and label_value.endswith(';'):
                        row = next(iterator)
                        row_list = row2list(row)
                        label_value = label_value + ' '.join([str(x) for x in row_list])

                    actions[action][label_name] = label_value
                    row = next(iterator)
                    row_list = row2list(row)

                    # check if a plate layout was added to the plate
                    if action == 'Plate' and row_list == []:
                        row = next(row_iterator)
                        row_list = row2list(row)
                        if row_list != []:
                            if row_list[0] == '<>':
                                subset = actions['Plate']['Plate area']
                                actions[action]['plate_layout'] = parse_data(excel_sheet, identifier='<>', subset=subset, plate=plate_format).rename({'value':'sample_type'}, axis=1)
                                # skip to end of plate layout
                                while row_list != []:
                                    row = next(iterator)
                                    row_list = row2list(row)
                        break

            return actions

        def format_config(row_list):
            '''formatting system config data from tecan file'''

            if len(row_list) == 1 and 'Method name:' in row_list[0]:
                key = 'Method'
                val = row_list[0].strip('Method name: ')
            elif 'Device:' in row_list[0] or 'Application:' in row_list[0]:
                key = row_list[0].split(': ')[0]
                val = ' '.join([row_list[0].split(': ')[1], row_list[1]])
            elif len(row_list) == 2:
                key = row_list[0]
                val = row_list[1]
            else:
                key = row_list[0]
                val = row_list[1:]

            return (key.strip(':'), val)
            
        wb_obj = openpyxl.load_workbook(excel_sheet)
        sheet_obj = wb_obj.active

        meta_data = dict()
        row_iterator = iter(sheet_obj.rows)
        config = True

        for row in row_iterator:
            row_list = row2list(row)
            
            if not row_list:
                # skip empty rows
                pass

            elif 'List of actions in this measurement script:' in row_list:
                config = False
                protocol = process_actions(row_iterator)
            
            elif config:
                key, val = format_config(row_list)
                meta_data[key] = val

        return meta_data, protocol


def parse_data(file: str, identifier: str, subset: str='A1-H12', label: str='value', skip: int=0, plate: int=96):
    """
    file: str, path to the tecan excel file
    identifier: str, identifier to find data
    subset: str, region of the excel sheet to use to retrieve data
    label: str, label of the data
    skip: int, number of appearances of identifier to skip
    plate: int, number of wells on plate

    Function to parse data from Tecan excel sheet
    """
    wb_obj = openpyxl.load_workbook(file)
    sheet_obj = wb_obj.active

    IMPLEMENTED = [96, 384]
    assert plate in IMPLEMENTED, f'plate format {plate} not implemented! {IMPLEMENTED}'

    # helper dicts for different plates
    usecols = {
        96: 'B:M',
        384: 'B:Y'
    }
    indices = {
        96: list('ABCDEFGH'),
        384: list('ABCDEFGHIJKLMNOP')
    }
    offsets = {
        96: 8,
        384: 16
    }

    count = 0
    for row in sheet_obj.iter_rows():
        cell = row[0]
        if cell.value == identifier and count == skip:
            if identifier != '<>':
                # TODO: fix this if there is a plate map, which has sample names (like PC) in the way
                start_row = cell.row + 1
            else:
                start_row = cell.row
            
            # locating end row
            end_row = start_row + offsets[plate]
            break

        elif cell.value == identifier:
            count += 1

    # read data from excel region
    skip = lambda x: x not in range(start_row - 1, end_row)
    df = pd.read_excel(file, usecols=usecols[plate], skiprows=skip)
    index = indices[plate]
    df.index = index
    df.columns = df.columns.astype(str)

    # select subsets
    subsets = subset.split(';')
    data = pd.DataFrame()

    for subset in subsets:
        start, end = subset.split('-')
        start_row = start[0]
        end_row = end[0]
        start_col = start[1:]
        end_col = end[1:]

        df_sub = df.loc[start_row : end_row, start_col : end_col]
        df_sub = df_sub.reset_index()
        df_melted = pd.melt(df_sub, id_vars='index', var_name='column', value_name=label)
        df_melted['well'] = df_melted['index'] + df_melted['column'].astype(str)
        df_melted.drop(['index', 'column'], axis=1, inplace=True)
        
        data = pd.concat([data, df_melted], ignore_index=True)
    
    data.drop_duplicates(subset='well', inplace=True)
    return data


def parse_wavel_scan(file: str, identifier: str, n_wavelengths: int, plate_region: str='A1-H12', label: str='value'):
    """
    file: str, path to the tecan excel file
    identifier: str, identifier to find data
    n_wavelengths: int, number of wavelengths measured
    plate_region: str, region of the plate with samples
    label: str, name of the measured variable

    Function to parse wavelength scan data from Tecan excel files
    """
    wb_obj = openpyxl.load_workbook(file)
    sheet_obj = wb_obj.active
    for row in sheet_obj.iter_rows():
        cell = row[0]
        if cell.value == identifier:
            excel_start_row = cell.row
            excel_end_row = int(excel_start_row + n_wavelengths)
            break

    subsets = plate_region.split(';')
    all_wells = []
    for subset in subsets:
        start, end = subset.split('-')
        start_row = start[0]
        end_row = end[0]
        start_col = start[1:]
        end_col = end[1:]

        rows = [chr(i) for i in range(ord(start_row),ord(end_row)+1)]
        cols = [i for i in range(int(start_col), int(end_col) + 1)]

        for x in cols:
            for y in rows:
                all_wells.append(y + str(x))
    
    end_col_letter = string.ascii_uppercase[len(all_wells)]
    skip = lambda x: x not in range(excel_start_row - 1, excel_end_row + 1)
    df = pd.read_excel(file, usecols=f'A:{end_col_letter}', skiprows=skip)
    df_melted = pd.melt(df, ignore_index=True, id_vars="Wavel.", var_name='Well', value_name=label)

    return df_melted